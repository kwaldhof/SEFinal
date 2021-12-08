import pandas as pd
from openpyxl import load_workbook
from pandas.core import series
import numpy as np
from pandas.core.arrays import sparse
def write():
    fname = "Tim_Test_data.xlsx"

    data = pd.read_excel(fname,"Sheet2") #read the excel file
    print(data)
    new_data = [[2,"WAC",3,3]] #Fetching new record from the frontend (right now I am just set up a record)
    df2 = pd.DataFrame(new_data, columns=['Team_ID',"Team_Name", "League_ID", "Player_Count"])# Convert this new record to dataframe
    data = data.append(df2,ignore_index=True)#append to the data we got from the excel
    print(data)

    book = load_workbook(fname)
    writer = pd.ExcelWriter(fname,engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)#trying to set up not overwrite the excel

    data.to_excel(writer,sheet_name="Sheet2",index=False)#push the data to the excel

    writer.save()

# League_Name,Organizer,Sport,Games
def create_league(League_Name,Org_Name_input,Sport_input):
    fname = "Tim_Test_data.xlsx"
    Sport = pd.read_excel (fname, 'Sport')
    Organizer = pd.read_excel (fname, 'Organizer')
    print(Organizer)
    for i in range(len(Sport)):
        if Sport_input == Sport.at[i,"Sport_Name"]:
            Sport_ID = Sport.at[i,"Sport_ID"]
    for j in range(len(Organizer)):
        if Org_Name_input == Organizer.at[j,"Org_Name"]:
            Organizer_ID = (Organizer.at[j,"Org_ID"])

    data = pd.read_excel(fname,"League") #read the excel file
    new_data = [[len(data),League_Name,Organizer_ID,Sport_ID,0]] #Fetching new record from the frontend (right now I am just set up a record)
    df2 = pd.DataFrame(new_data, columns=['League_ID',"League_Name","Organizer","Sport", "Games"])# Convert this new record to dataframe
    data = data.append(df2,ignore_index=True)#append to the data we got from the excel
    print(data)

    book = load_workbook(fname)
    writer = pd.ExcelWriter(fname,engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)#trying to set up not overwrite the excel

    data.to_excel(writer,sheet_name="League",index=False)#push the data to the excel

    writer.save()


create_league("WACCC","Timmm","Soccer")